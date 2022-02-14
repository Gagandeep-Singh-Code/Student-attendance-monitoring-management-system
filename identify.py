from global_variables import PERSON_GROUP_ID
import os, urllib
import pyodbc
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, cell, column_index_from_string
import time
from azure.cognitiveservices.vision.face import FaceClient
from msrest.authentication import CognitiveServicesCredentials

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports.xlsx")
sheet = wb['attendance']

def getDateColumn():
	for i in range(1, sheet.max_column + 1):
		col = get_column_letter(i)
		if sheet['%s%s'% (col,'1')].value == currentDate:
			return col


KEY = 'ENTER API KEY HERE'


 
ENDPOINT = "ENTER BASE URL HERE"

face_client = FaceClient(ENDPOINT, CognitiveServicesCredentials(KEY))

#connect = connect = sqlite3.connect("Face-DataBase")
server = 'ENTER SERVER NAME HERE'
database = 'ENTER DATABASE NAME HERE'
username = 'ENTER USERNAME HERE'
password = 'ENTER PASSWORD HERE'
connect = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
c = connect.cursor()


attend = [0 for i in range(200)]

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'cropped')
for filename in os.listdir(directory):
	if filename.endswith(".jpg"):
		image = (os.path.join(directory, filename))
		img = open(image, 'r+b')
		res = face_client.face.detect_with_stream(img, detectionModel='detection_03')
		time.sleep(6)
		img.close()
		if len(res) != 1:
			print ("No face detected.")
			continue

		faceIds = []
		for face in res:
			faceIds.append(face.face_id)
		
		res = face_client.face.identify(faceIds, PERSON_GROUP_ID)
		#print (filename)
		#print (res)
		for face in res:
			if not face.candidates:
				print ("Unknown")
			else:
				personId = face.candidates[0].person_id
				#print(personId)
				c.execute("SELECT * FROM Students WHERE personID = ?", (personId))
				row = c.fetchone()
				#print (row)
				attend[int(row[0])] += 1
				print (row[1] + " recognized")


for row in range(2, sheet.max_row + 1):
	rn = sheet['A%s'% row].value
	if rn is not None:
		rn = rn[1:]
		if attend[int(rn)] != 0:
			col = getDateColumn()
			sheet['%s%s' % (col, str(row))] = 1

wb.save(filename = "reports.xlsx")

import shutil
folder = './cropped'
for the_file in os.listdir(folder):
    file_path = os.path.join(folder, the_file)
    try:
        if os.path.isfile(file_path):
            os.unlink(file_path)
        #elif os.path.isdir(file_path): shutil.rmtree(file_path)
    except Exception as e:
        print(e)
